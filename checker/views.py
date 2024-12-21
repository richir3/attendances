import json, random, openpyxl, datetime, base64, qrcode
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
from .models import *
from .forms import AttenderForm, EventForm, AddEventForm, EmailForm
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.urls import reverse, reverse_lazy
from django.core.mail import EmailMultiAlternatives
from django.conf import settings
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from rest_framework import generics, permissions
from .serializers import *
from django.views.generic.edit import UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin

# Create your views here.
@login_required
def home(request):
    class Functionality:
        def __init__(self, name, url):
            self.name = name
            self.url = url

        def __repr__(self):
            return f"{self.name.lower().replace(' ', '-')}"

    functionalities = [
        Functionality("Lector QR", reverse("qr_reader_view")),
        Functionality("Gestionar asistentes", reverse("list_attenders")),
        Functionality("Gestionar eventos", reverse("add_event")),
    ]

    return render(request, "home.html", {"functionalities": functionalities})

@login_required
def qr_reader_view(request):
    codes = dict()
    for a in Attender.objects.all():
        codes[a.code] = a.name + " " + a.surname

    form = EventForm()
    return render(request, "reader.html", {"form": form, "codes": codes})

@login_required
def post_qr_data(request):
    if request.method == "POST":
        data = json.loads(request.body)
        try:
            attender = Attender.objects.get(code=data['content'])  
            event = Event.objects.get(id=data['date'])
            # come faccio a sapere se Attendance.objects.get(event=event, user=attender) esiste nel mio db?
            if Attendance.objects.filter(event=event, user=attender).exists():
                response = JsonResponse({'status': 'error', 'error': 'El asistente ya asistió al evento.'}, status=400)
            else:
                attendance = Attendance(event=event, user=attender)
                
                # # send an email to the attender to confirm the attendance
                # context = {
                #     "event": event,
                # }

                # html_content = render_to_string("email/confirm_attendance.html", context=context)
                # text_content = strip_tags(html_content)

                # email = EmailMultiAlternatives(
                #     subject="Asistencia confirmada",
                #     body=text_content,
                #     from_email=f'"Organización de eventos" <{settings.EMAIL_HOST_USER}>',
                #     to=[attender.email],
                # )

                # email.attach_alternative(html_content, "text/html")

                # try:
                #     email.send()
                # except Exception as e:
                #     response = JsonResponse({'status': 'error con el correo', 'error': str(e)}, status=500)

                attendance.save()
                response = JsonResponse({'status': 'success', 'message': 'Asistencia guardada'}, status=200)
        except KeyError as e:
            response = JsonResponse({'status': 'error', 'error': 'Invalid data'}, status=400)
        except Attender.DoesNotExist as e:
            response = JsonResponse({'status': 'error', 'error': 'Asistente no encontrado'}, status=404)
        except Event.DoesNotExist as e:
            response = JsonResponse({'status': 'error', 'error': 'Evento no encontrado'}, status=404)
        except json.JSONDecodeError as e:
            response = JsonResponse({'status': 'error', 'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            response = JsonResponse({'status': 'error', 'error': str(e)}, status=500)
        finally:
            return response
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@login_required
def list_attenders(request):    
    class AttenderShow:
        def __init__(self, attender):
            self.attender = attender
            self.attendances = Attendance.objects.filter(user=attender).count()
        
        def __repr__(self):
            return f"{self.attender.name} {self.attender.surname} - {self.attendances}"
    
    attenders = [AttenderShow(a) for a in Attender.objects.all().order_by("name")]

    context = {
        "attenders": attenders,
    }
    return render(request, "attenders.html", context=context)

@login_required
def add_attender(request):
    if request.method == "POST":
        form = AttenderForm(request.POST)
        if form.is_valid():
            form.save()
            attender = Attender.objects.get(name=form.cleaned_data['name'], surname=form.cleaned_data['surname'])
            try:
                send_qr(attender.id)
            except Exception as e:
                pass
            
            return redirect("list_attenders")
    else:
        form = AttenderForm()
        return render(request, "add_attender.html", {"form": form})

@login_required
def attender_overview(request, pk):
    attender = Attender.objects.get(pk=pk)
    context = {
        "attender": attender,
        "events" : Event.objects.all().order_by("date"),
        "attendances" : [a.event for a in Attendance.objects.all().filter(user=attender)]
    }

    return render(request, "attender.html", context=context)

# modify or delete attender
class AttenderUpdateView(LoginRequiredMixin, UpdateView):
    model = Attender
    form_class = AttenderForm
    template_name = 'attender_form.html'
    success_url = reverse_lazy('list_attenders')  # Cambia con il nome della tua lista di attenders

# Elimina un Attender
class AttenderDeleteView(LoginRequiredMixin, DeleteView):
    model = Attender
    template_name = 'attender_confirm_delete.html'
    success_url = reverse_lazy('list_attenders')

@login_required
def download_attendances(request):
    NO = "NO"
    YES = "SI"
    events = Event.objects.all().order_by("date")
    event_dates = [e.date.strftime("%d/%m/%Y") for e in events]

    attenders = Attender.objects.all()

    data = {}
    for a in attenders:
        data[a] = []
        for e in events:
            #if a.events.filter(pk=e.pk).exists():
            if Attendance.objects.filter(event=e, user=a).exists():
                data[a].append(YES)
            else:
                data[a].append(NO)

    df = pd.DataFrame(data).T
    df.columns = event_dates

    # insert a column in the dataframe with the total attendances, in the second position
    df.insert(0, "Total", df.apply(lambda x: x.value_counts().get(YES, 0), axis=1))

    df.insert(0, "Cofradia", df.apply(lambda x: x.name.brotherhood.name, axis=1))

    # Creazione di un buffer in memoria
    buffer = BytesIO()

    # Scrittura iniziale del DataFrame in un file Excel
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=True, sheet_name="Attendances")

    # Caricamento del file Excel in openpyxl per le modifiche
    buffer.seek(0)  # Riavvolgi il buffer per leggerlo da openpyxl
    wb = openpyxl.load_workbook(buffer)
    ws = wb.active  # Seleziona il primo foglio

    ## Modifiche al foglio di lavoro
    # imposta la prima riga come tipo data
    for cell in ws[1][1:]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = "DD/MM/YYYY"
    
    # imposta le celle con YES in verde e quelle con NO in rosso
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.value == YES:
                cell.fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif cell.value == NO:
                cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # adatta la larghezza delle colonne
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Scrivi le modifiche nel buffer
    new_buffer = BytesIO()  # Nuovo buffer per la risposta finale
    wb.save(new_buffer)
    new_buffer.seek(0)

    # Configura la risposta HTTP per il download del file
    response = HttpResponse(
        new_buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    datefile = datetime.datetime.now().strftime("%d-%m-%Y--%H:%M")

    response["Content-Disposition"] = f'attachment; filename="asistencia_{datefile}.xlsx"'
    return response

@login_required
def add_event(request):
    if request.method == "POST":
        form = AddEventForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("home")
    else:
        form = AddEventForm()
        context = {
            "form": form,
            "events": Event.objects.filter(date__gte=datetime.date.today())
        }
        return render(request, "add_event.html", context)

@login_required
def send_qr_code_mail(request, attender_id):
    if request.method == "POST":
        try:
            send_qr(attender_id)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
        return JsonResponse({"status": "success", "message": "Correo enviado"}, status=200)
    else:
        return JsonResponse({"status": "error", "message": "Invalid request method"}, status=405)

# def send_qr_image(attender_id):
#     attender = Attender.objects.get(id=attender_id)

#     # generate qr
#     code = attender.code
#     qr = qrcode.QRCode(
#         version=1,
#         error_correction=qrcode.constants.ERROR_CORRECT_L,
#         box_size=10,
#         border=4,
#     )
#     qr.add_data(code)
#     qr.make(fit=True)

#     # Converti il QR Code in un'immagine PNG
#     img = qr.make_image(fill_color="black", back_color="white")
#     buffer = BytesIO()
#     img.save(buffer, format="PNG")
#     buffer.seek(0)

#     # Codifica l'immagine in base64
#     img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

#     context = {
#         "attender": attender,
#         "qr": img_base64,
#     }


#     # Render del template
#     html_content = render_to_string("email/qr_email.html", context=context)
#     text_content = strip_tags(html_content)
    
#     # Configura la mail
#     email = EmailMultiAlternatives(
#         subject="Tu código QR para el evento",
#         body=text_content,
#         from_email=f'"Organización de eventos" <{settings.EMAIL_HOST_USER}>',
#         to=[attender.email],
#     )

#     email.attach_alternative(html_content, "text/html")
#     email.attach(f"codigo qr {attender.name} {attender.surname}.png", buffer.getvalue(), "image/png")
    
#     email.send()

def send_qr_code_mail(request, attender_id):
    attender = Attender.objects.get(id=attender_id)
    if request.method == "POST":
        form = EmailForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data["email"]
            send_qr(attender_id, email)
            return redirect("list_attenders")
    else:
        form = EmailForm()
        context = {
            "attender": attender,
            "form": form,
        }
        return render(request, "mail_qr_to.html", context=context)


# Send PDF with information and QR code
def send_qr(attender_id, emailaddress = None):
    attender = Attender.objects.get(id=attender_id)
    if emailaddress == None:
        emailaddress = attender.brotherhood.email
    # Strings
    SUBJECT = f"Código QR de {attender.name.upper()} {attender.surname.upper()}"
    FROM_EMAIL = f'"Organización Pregón 2025" <{settings.EMAIL_HOST_USER}>'
    INFORMATION = "Información del asistente"
    NAME = f"Nombre: {attender.name.upper()}"
    SURNAME = f"Apellido: {attender.surname.upper()}"
    BROTHERHOOD = f"Cófradia: {attender.brotherhood.name.upper()}"
    QR_CODE = f"Código QR:"

    # Generate qr code
    code = attender.code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(code)
    qr.make(fit=True)

    qr_img = qr.make_image(fill_color="black", back_color="white")
    qr_buffer = BytesIO()
    qr_img.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)

    # Create PDF
    pdf_buffer = BytesIO()
    pdf = canvas.Canvas(pdf_buffer, pagesize=letter)
    pdf.setFont("Helvetica-Bold", 20)
    pdf.drawCentredString(300, 750, INFORMATION)
    pdf.setFont("Helvetica", 18)
    pdf.drawCentredString(300, 710, NAME)
    pdf.drawCentredString(300, 690, SURNAME)
    pdf.drawCentredString(300, 670, BROTHERHOOD)
    pdf.drawCentredString(300, 650, QR_CODE)

    # QR code to PDF
    qr_img_reader = ImageReader(qr_buffer)
    pdf.drawImage(qr_img_reader, 150, 300, width=300, height=300)

    # Close PDF
    pdf.showPage()
    pdf.save()
    pdf_buffer.seek(0)

    # Send email
    context = {
        "attender": attender,
    }

    if emailaddress is not None:
        context["name"] = attender.name

    html_content = render_to_string("email/qr_email.html", context=context)
    text_content = strip_tags(html_content)

    email = EmailMultiAlternatives(
        subject=SUBJECT,
        body=text_content,
        from_email=FROM_EMAIL,
        to=[emailaddress],
    )

    email.attach_alternative(html_content, "text/html")
    email.attach(f"{attender.name}_{attender.surname}_QR.pdf", pdf_buffer.getvalue(), "application/pdf")

    email.send()

# API REST
class AttenderListCreate(generics.ListCreateAPIView):
    queryset = Attender.objects.all()
    serializer_class = AttenderSerializer
    permission_classes = [permissions.IsAdminUser]

    # how can i execcute a function after the creation of an object?
    def perform_create(self, serializer):
        instance = serializer.save()
        try:
            send_qr(instance.id)
        except Exception as e:
            pass

class AttenderRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Attender.objects.all()
    serializer_class = AttenderSerializer
    permission_classes = [permissions.IsAdminUser]

class EventListCreate(generics.ListCreateAPIView):
    queryset = Event.objects.all()
    serializer_class = EventSerializer
    permission_classes = [permissions.IsAdminUser]

class EventRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Event.objects.all()
    serializer_class = EventSerializer
    permission_classes = [permissions.IsAdminUser]

class AttendanceListCreate(generics.ListCreateAPIView):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [permissions.IsAdminUser]

class AttendanceRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [permissions.IsAdminUser]

class BrotherhoodListCreate(generics.ListCreateAPIView):
    queryset = Brotherhood.objects.all()
    serializer_class = BrotherhoodSerializer
    permission_classes = [permissions.IsAdminUser]

class BrotherhoodRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Brotherhood.objects.all()
    serializer_class = BrotherhoodSerializer
    permission_classes = [permissions.IsAdminUser]